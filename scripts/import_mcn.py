#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MCN 刊例导入脚本（小红书表）
读取 MCN 的 Excel 刊例 → 清洗映射 → 生成 data/data-full.{json,js} 与 data/insights.{json,js}

用法：
  python3 scripts/import_mcn.py "<刊例.xlsx路径>" ["sheet名"] ["MCN名称"]

注意：下方 INDUSTRIES / PERSONA / DIM / TIER / CROWD_LIST / AGE_LIST 必须与 js/config.js 保持一致。
"""
import sys, json, math
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl，请先安装：pip install openpyxl")
    sys.exit(1)

# ===== 与 js/config.js 保持一致的口径 =====
INDUSTRIES = ["美妆个护", "时尚穿搭", "母婴亲子", "食品饮料", "家居家装", "运动户外", "3C数码",
              "汽车出行", "旅行", "萌宠", "教育知识", "医疗健康", "金融财经", "游戏", "本地生活", "职场成长"]
IDENTITY = ["素人", "达人", "MCN机构", "品牌合作人"]
PERSONA = ["专业测评官", "生活方式种草家", "审美穿搭官", "干货知识博主", "剧情演绎博主", "高互动挑战博主"]
DIM = ["测评种草", "教程攻略", "好物合集", "探店体验", "剧情演绎", "专业科普"]
TIER = ["<1W", "1-5W", "5-10W", "10-50W", "50-100W", "100-500W", "500W+"]
CROWD_LIST = ["时尚潮流", "精致妈妈", "都市白领", "Z世代", "资深中产", "学生党", "小镇青年", "银发一族"]
AGE_LIST = ["18以下", "18-24", "25-34", "35-44", "45+"]

KEYS = ["id", "name", "ide", "mcn", "per", "dim", "tier", "tag", "vmask",
        "male", "female", "age", "city", "crowd", "ios",
        "p_img", "p_video", "c_img", "c_video", "eng", "notes30", "fans", "home_url", "pgy_url"]

# MCN 的「达人类型」→ (行业, 标签)
CATEGORY_MAP = {
    "时尚": ("时尚穿搭", "时尚"), "穿搭": ("时尚穿搭", "穿搭"),
    "美妆": ("美妆个护", "美妆"), "护肤": ("美妆个护", "护肤"),
    "母婴亲子": ("母婴亲子", "亲子"), "育儿": ("母婴亲子", "育儿"),
    "美食": ("食品饮料", "美食"), "食品": ("食品饮料", "食品"),
    "家居家装": ("家居家装", "家居"), "家居家电": ("家居家装", "家居"), "家居": ("家居家装", "家居"), "家电": ("家居家装", "家居"),
    "数码": ("3C数码", "数码"), "科技数码": ("3C数码", "数码"), "科技": ("3C数码", "科技"),
    "AI": ("3C数码", "AI"), "AI应用": ("3C数码", "AI"), "AIGC": ("3C数码", "AI"),
    "汽车": ("汽车出行", "汽车"), "出行": ("汽车出行", "汽车"),
    "旅行": ("旅行", "旅行"), "旅游": ("旅行", "旅行"),
    "财经": ("金融财经", "财经"), "财经科普": ("金融财经", "财经"), "财经投资": ("金融财经", "财经"), "金融": ("金融财经", "财经"),
    "教育": ("教育知识", "教育"), "教育培训": ("教育知识", "教育"), "知识科普": ("教育知识", "知识"), "科普": ("教育知识", "知识"),
    "游戏": ("游戏", "游戏"), "萌宠": ("萌宠", "萌宠"), "宠物": ("萌宠", "宠物"),
    "运动": ("运动户外", "运动"), "健身": ("运动户外", "健身"), "户外": ("运动户外", "户外"),
    "健康": ("医疗健康", "健康"), "医疗": ("医疗健康", "医疗"),
    "职场": ("职场成长", "职场"), "本地生活": ("本地生活", "本地生活"), "探店": ("本地生活", "探店"),
}

def contains_any(text, keys):
    return any(k in text for k in keys)

def persona_of(text):
    if contains_any(text, ["测评", "评测", "实测", "拆解", "数码", "科技", "AI", "汽车", "家电"]):
        return "专业测评官"
    if contains_any(text, ["时尚", "美妆", "护肤", "穿搭", "明星穿搭"]):
        return "审美穿搭官"
    if contains_any(text, ["知识", "科普", "教育", "财经", "培训", "解读", "分析", "洞察"]):
        return "干货知识博主"
    if contains_any(text, ["剧情", "搞笑", "演绎"]):
        return "剧情演绎博主"
    return "生活方式种草家"

def dim_of(text):
    if contains_any(text, ["测评", "评测", "实测", "拆解", "横评"]):
        return "测评种草"
    if contains_any(text, ["教学", "教程", "培训", "软件教学", "攻略", "解读"]):
        return "教程攻略"
    if contains_any(text, ["探店"]):
        return "探店体验"
    if contains_any(text, ["剧情", "搞笑", "演绎"]):
        return "剧情演绎"
    if contains_any(text, ["科普", "知识", "分析", "洞察", "解读"]):
        return "专业科普"
    return "好物合集"

def tier_of(fans):
    if fans < 10000: return "<1W"
    if fans < 50000: return "1-5W"
    if fans < 100000: return "5-10W"
    if fans < 500000: return "10-50W"
    if fans < 1000000: return "50-100W"
    if fans < 5000000: return "100-500W"
    return "500W+"

def parse_categories(raw):
    return [t for t in str(raw or "").replace("、", ",").replace("/", ",").split(",") if t.strip()]

def to_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip().replace(",", "").replace("w", "").replace("W", "").replace("万", "")
    if not t or t in ("/", "——", "-", "面议", "待定"):
        return 0.0
    try:
        return float(t)
    except Exception:
        return 0.0

def to_int(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    t = str(v).strip().replace(",", "").replace("￥", "").replace("元", "")
    if not t or t in ("/", "——", "-", "面议", "待定"):
        return 0
    try:
        return int(float(t))
    except Exception:
        return 0

def build_record(row, mcn_name):
    _, name, cat_raw, desc_raw, uid, fans_wan, city, p_img, p_video, home_url = row[:10]
    if not name or not home_url or not str(home_url).startswith("http"):
        return None
    text = f"{cat_raw or ''} {desc_raw or ''}"

    cats = parse_categories(cat_raw)
    bits = 0
    mapped_tags = []
    for c in cats:
        hit = CATEGORY_MAP.get(c.strip())
        if hit:
            ind = INDUSTRIES.index(hit[0])
            bits |= (1 << ind)
            mapped_tags.append(hit[1])
    vmask = bits
    tag = ",".join(mapped_tags) or str(cat_raw or "").strip()

    fans = int(to_float(fans_wan) * 10000)  # 刊例 F 列单位是「万」

    p_img = to_int(p_img)
    p_video = to_int(p_video)
    # 画像/互动缺失 → 0；CPE 用「报价 ÷ 预估互动量」做参考估值（后续接蒲公英真实数据再替换）
    interactions = max(1, int(fans * 0.03))
    c_img = int(round(p_img / interactions)) if p_img else 0
    c_video = int(round(p_video / interactions)) if p_video else 0

    city_flat = [str(city).strip(), 1000] if city and str(city).strip() else []

    return [
        str(uid or ""),                 # id（小红书账号ID）
        str(name or ""),                # name
        2,                              # ide = MCN机构
        mcn_name,                       # mcn
        PERSONA.index(persona_of(text)),# per
        DIM.index(dim_of(text)),        # dim
        TIER.index(tier_of(fans)),      # tier
        tag,                            # tag
        vmask,                          # vmask
        0, 0,                           # male, female（待补）
        [0, 0, 0, 0, 0],                # age（待补）
        city_flat,                      # city
        [0, 0, 0, 0, 0, 0, 0, 0],       # crowd（待补）
        0,                              # ios（待补）
        p_img, p_video, c_img, c_video, # 报价 + CPE
        0, 0,                           # eng, notes30（待补）
        fans,                           # fans
        str(home_url or ""),            # home_url
        "",                             # pgy_url（待补）
    ]

def build_insights(total):
    out = {}
    for ind in INDUSTRIES:
        out[ind] = {
            "hot_trends": [f"{ind}内容在小红书持续走热，真实体验与干货清单更容易出圈。"],
            "topic_ideas": [f"{ind}好物实测合集", f"{ind}新手避坑指南", f"{ind}高性价比清单"],
            "strategy": "头部专业测评建立信任 + 腰部真实种草 + 尾部挑战/剧情拉高传播，建议 1:3:6 比例投放。",
        }
    out["__all__"] = {
        "hot_trends": ["小红书当前热点集中在真实体验、干货清单与沉浸式挑战三类内容。"],
        "topic_ideas": ["跨品类「真实体验 + 前后对比」选题", "高互动「挑战/测评」内容二创"],
        "strategy": "按行业选择垂类博主，优先覆盖目标城市与预算区间，图文+视频组合投放。",
    }
    return out

def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "/Users/xiaomadaidaidemacbookpro/Desktop/依依【乾派文化】2026年8月刊例.xlsx"
    sheet = sys.argv[2] if len(sys.argv) > 2 else "小红书"
    mcn = sys.argv[3] if len(sys.argv) > 3 else "乾派文化"

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        print(f"未找到 sheet「{sheet}」，可用：{wb.sheetnames}")
        sys.exit(1)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    records = []
    seen = set()
    for r in rows[2:]:  # 跳过图片行 + 表头行
        rec = build_record(r, mcn)
        if not rec:
            continue
        key = rec[KEYS.index("home_url")]
        if key in seen:
            continue
        seen.add(key)
        records.append(rec)

    meta = {
        "keys": KEYS,
        "industries": INDUSTRIES,
        "identity": IDENTITY,
        "persona": PERSONA,
        "dim": DIM,
        "tier": TIER,
        "crowd_list": CROWD_LIST,
        "age_list": AGE_LIST,
        "unit": "per-mille（除以 1000 即比例）",
        "total": len(records),
        "version": "2026-08-30-mcn",
        "generated_at": "2026-08-30",
        "source": f"MCN刊例·{mcn}·{sheet}",
    }
    data = {"meta": meta, "total": len(records), "records": records}
    insights = build_insights(len(records))

    import os
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")
    for fname, obj in [("data-full.json", data), ("insights.json", insights)]:
        with open(os.path.join(out_dir, fname), "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, separators=(",", ":"))
    with open(os.path.join(out_dir, "data-full.js"), "w", encoding="utf-8") as f:
        f.write("window.PGW_DATA_FULL = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n")
    with open(os.path.join(out_dir, "insights.js"), "w", encoding="utf-8") as f:
        f.write("window.PGW_INSIGHTS = " + json.dumps(insights, ensure_ascii=False, separators=(",", ":")) + ";\n")

    print(f"✔ 导入成功：{sheet} 表 → {len(records)} 位博主（MCN：{mcn}）")
    print(f"✔ 已生成 data/data-full.json / data-full.js / insights.json / insights.js")

if __name__ == "__main__":
    main()
